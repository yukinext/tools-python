#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jan 14 07:47:11 2019

@author: yuki_next
"""

import requests
from bs4 import BeautifulSoup
import argparse
import openpyxl
import datetime
import os
import pickle
import logging
import logging.config
import re
import sys

DEFAULT_URL_ROOT = "https://www.tv-asahi.co.jp/goodmorning/uranai/"
MAP_PAGE_CONS_ID = {
        "#ohitsuji":0,
        "#ousi":1,
        "#futago":2,
        "#kani":3,
        "#sisi":4,
        "#otome":5,
        "#tenbin":6,
        "#sasori":7,
        "#ite":8,
        "#yagi":9,
        "#mizugame":10,
        "#uo":11,        
        }

MAP_PAGE_CONS = {
        0:"おひつじ座",
        1:"おうし座",
        2:"ふたご座",
        3:"かに座",
        4:"しし座",
        5:"おとめ座",
        6:"てんびん座",
        7:"さそり座",
        8:"いて座",
        9:"やぎ座",
        10:"みずがめ座",
        11:"うお座",
        }

logging.config.dictConfig({
    'version': 1,
    'formatters': {
        'myFormatter': {
            'format': '%(asctime)s:%(message)s',
        }
    },
    'handlers': {
        'default': {
            'class': 'logging.StreamHandler',
            'level': 'DEBUG',
            'stream': 'ext://sys.stdout',
            'formatter': 'myFormatter',
        }
    },
    'root': {
        'level': 'DEBUG',
        'handlers': ['default'],
    },
    'disable_existing_loggers': False,
})

logger = logging.getLogger()

class ConstellationInfo(object):
    def __init__(self, cons_id, cons_id_name, name, rank):
        self.cons_id = cons_id
        self.cons_id_name = cons_id_name
        self.name = name
        self.rank = rank
        self.advice = ""
        self.lucky_color = None
        self.lucky_item = None
        self.fortune_economic = None
        self.fortune_love = None
        self.fortune_work = None
        self.fortune_health = None

    def __repr__(self):
        items = ("%s = %r" % (k, v) for k, v in self.__dict__.items())
        return "<%s: {%s}>" % (self.__class__.__name__, ', '.join(items))  

def get_target_date(bs):
    t = bs.select_one("#ranking").p.get_text().strip()
    m = re.match(r"(?P<month>\d+)\D+(?P<day>\d+)\D+.*", t)
    today = datetime.date.today()
    
    if m:
        ret = today.replace(month=int(m["month"]), day=int(m["day"]))
        if 1 < abs((today - ret).days):
            ret = ret.replace(year=ret.year - 1)
    
    return ret

def main():
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--output_dir", default="out")
    parser.add_argument("--url_root", default=DEFAULT_URL_ROOT)
    parser.add_argument("--output_fn", default="_Constellation.xlsx")
    parser.add_argument("--sn_rank", default="ranking")
    parser.add_argument("--sn_info", default="info")
    parser.add_argument("--oauth_json_filename", 
                        default=os.path.join(
                                os.path.dirname(os.path.abspath(sys.argv[0])),
                                "cred.json"))

    args = parser.parse_args()
    infos = dict()
    
    # target_date = datetime.date.today()
    
    res = requests.get(args.url_root)
    bs = BeautifulSoup(res.content, features="html.parser")
    
    target_date = get_target_date(bs)
    
    for i, a in enumerate(bs.ul.find_all("a")):
        cons_id_name = a["href"]
        cons_id = MAP_PAGE_CONS_ID[cons_id_name]
        infos[cons_id] = (ConstellationInfo(cons_id, cons_id_name, MAP_PAGE_CONS[cons_id], i + 1))
 
    for info in infos.values():
        soup = bs.select_one(info.cons_id_name)
        read_area = soup.find("div", "read-area")
        info.advice = read_area.p.string.strip()
        _, info.lucky_color = read_area.select_one(".lucky-color-txt").next_sibling.strip().split("：")
        _, info.lucky_item = read_area.select_one(".key-txt").next_sibling.strip().split("：")
        info.fortune_economic = len(soup.select_one(".lucky-money").select(".icon-money"))   
        info.fortune_love = len(soup.select_one(".lucky-love").select(".icon-love"))   
        info.fortune_work = len(soup.select_one(".lucky-work").select(".icon-work"))   
        info.fortune_health = len(soup.select_one(".lucky-health").select(".icon-health"))   
    
        logger.info(info)
    
    cache_filename = os.path.join(args.output_dir, target_date.strftime("%Y-%m-%d"))
    if not os.path.exists(args.output_dir):
        os.makedirs(args.output_dir)
        
    with open(cache_filename, "wb") as fp:
        pickle.dump(infos, fp)
        logger.info("save: %s" % cache_filename)
    
    output_filename = os.path.join(args.output_dir, args.output_fn)
    wb = None
    if os.path.exists(output_filename):
        logger.info("overwrite")
        wb = openpyxl.load_workbook(output_filename)

        if not args.sn_rank in wb:
            ws = wb.create_sheet(args.sn_rank)
            ws.append(
                    ["date"]
                    + [s for id, s in sorted(MAP_PAGE_CONS.items())]
                    )

        if not args.sn_info in wb:
            ws = wb.create_sheet(args.sn_info)
            ws.append(
                    ["date(金運/恋愛運/仕事運/健康運)"]
                    + [s for id, s in sorted(MAP_PAGE_CONS.items())]
                    )
    else:
        logger.info("create new file")
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = args.sn_rank
        ws.append(
                ["date"]
                + [s for id, s in sorted(MAP_PAGE_CONS.items())]
                )
    
        ws = wb.create_sheet(args.sn_info)
        ws.append(
                ["date(金運/恋愛運/仕事運/健康運)"]
                + [s for id, s in sorted(MAP_PAGE_CONS.items())]
                )

    ws = wb[args.sn_rank]
    row = ws[ws.max_row]
    if 1 < ws.max_row and row[0].value.date() == target_date:
        logger.info("[%s] exists %s" % (ws.title, target_date))
    else:
        logger.info("[%s] append %s" % (ws.title, target_date))
        ws.append(
                [target_date]
                + [info.rank for id, info in sorted(infos.items())]
                )
        

    ws = wb[args.sn_info]
    row = ws[ws.max_row]
    if 1 < ws.max_row and row[0].value.date() == target_date:
        logger.info("[%s] exists %s" % (ws.title, target_date))
    else:
        logger.info("[%s] append %s" % (ws.title, target_date))
        ws.append(
                [target_date]
                + ["%s,%s\n%d/%d/%d/%d\n%s" % (
                        info.lucky_color, info.lucky_item, 
                        info.fortune_economic, info.fortune_love, info.fortune_work, info.fortune_health,
                        info.advice) for id, info in sorted(infos.items())]
                    )
        
    wb.save(output_filename)
    
    if os.path.exists(args.oauth_json_filename):
        from oauth2client.service_account import ServiceAccountCredentials
        import gspread
        scope = ['https://spreadsheets.google.com/feeds',
                 'https://www.googleapis.com/auth/drive']
        credentials = ServiceAccountCredentials.from_json_keyfile_name(args.oauth_json_filename, scope)
        gc = gspread.authorize(credentials)
        gwb = gc.open(os.path.splitext(args.output_fn)[0])

        gws = gwb.worksheet(args.sn_rank)
        if 0 == gws.row_count:
            gws.append_row(
                    ["date"]
                    + [s for id, s in sorted(MAP_PAGE_CONS.items())]
                    )
        row = gws.row_values(gws.row_count)
        target_date = target_date.strftime("%Y-%m-%d")
        if 1 < gws.row_count and row[0] == target_date:
            logger.info("g[%s] exists %s" % (gws.title, target_date))
        else:
            logger.info("g[%s] append %s" % (gws.title, target_date))
            gws.append_row(
                    [target_date]
                    + [info.rank for id, info in sorted(infos.items())]
                    )

        gws = gwb.worksheet(args.sn_info)
        if 0 == gws.row_count:
            gws.append_row(
                    ["date(金運/恋愛運/仕事運/健康運)"]
                    + [s for id, s in sorted(MAP_PAGE_CONS.items())]
                    )
        row = gws.row_values(gws.row_count)
        if 1 < gws.row_count and row[0] == target_date:
            logger.info("g[%s] exists %s" % (gws.title, target_date))
        else:
            logger.info("g[%s] append %s" % (gws.title, target_date))
            gws.append_row(
                    [target_date]
                    + ["%s,%s\n%d/%d/%d/%d\n%s" % (
                            info.lucky_color, info.lucky_item, 
                            info.fortune_economic, info.fortune_love, info.fortune_work, info.fortune_health,
                            info.advice) for id, info in sorted(infos.items())]
                    )
    
if __name__ == "__main__":
    main()